#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@File    :   use_excel.py    
@Contact :   Johnd0712@hotmail.com
@License :   (C)Copyright 2017-2018, Krynn.cn

@Modify Time         @Author      @Version    @Desciption
-----------------    ---------    --------    -----------
2019-03-17 16:28      Raistlind    1.0         None
"""

# import lib
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors
import MySQLdb


class ExcelUtils():

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws_two = self.wb.create_sheet('第一个表单')
        self.ws.title = '0位表单'
        self.ws.sheet_properties.tabColor = 'ff0000'
        self.ws_three = self.wb.create_sheet()

    def do_sth(self):
        self.ws['A1'] = 66
        self.ws['A2'] = '测试'
        self.ws['A3'] = datetime.now()

        for row in self.ws_two['A1:E5']:
            for cell in row:
                cell.value = 2

        # 对数据进行求和
        self.ws_two['G1'] = '=SUM(A1:E1)'

        # 插入图片
        img = Image('./static/130402161P2_300.jpg')
        self.ws.add_image(img, 'B1')

        # 合并单元格
        self.ws.merge_cells('A4:E5')
        self.ws.unmerge_cells('A4:E5')

        # 设置文字
        font = Font(sz=18, color=colors.RED)
        self.ws['A2'].font = font

        # 保存
        self.wb.save('./static/test.xlsx')

    def read_xls(self):
        ws = load_workbook('./static/template.xlsx')
        names = ws.get_sheet_names()
        print(names)
        # wb = ws.active
        wb = ws['北京大学统计']
        for (i, row) in enumerate(wb.rows):
            if i < 2:
                continue
            year = wb['A{0}'.format(i+1)].value
            max_score = wb['B{0}'.format(i+1)].value
            avg_score = wb['C{0}'.format(i+1)].value

            print(year, max_score, avg_score)

            if year is None:
                continue

            conn = client.get_conn()
            cursor = conn.cursor()
            sql = 'insert into user_score (year, max_score, avg_score) values ({year}, {max_score} , {avg_score})'\
                .format(year=year, max_score=max_score, avg_score=avg_score)
            result = cursor.execute(sql)
            conn.autocommit(True)
            print(result)



    def get_conn(self):
        conn = MySQLdb.connect(
            db='user_grade',
            host='localhost',
            user='root',
            password='123456',
            charset='utf8',
            unix_socket='/Applications/MAMP/tmp/mysql/mysql.sock'
        )
        return conn


if __name__ == '__main__':
    client = ExcelUtils()
    # client.do_sth()
    client.read_xls()

