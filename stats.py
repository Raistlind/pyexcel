import MySQLdb
from openpyxl import load_workbook, Workbook
from openpyxl.chart import AreaChart, Reference


class GaokaoExport():

    def __init__(self):
        self.wb = load_workbook('./static/tmpl.xlsx')
        self.ws = self.wb.active
        self.ws.title = '成绩统计'
        self.ws.sheet_properties.tabColor = 'ff0000'

    def get_conn(self):
        try:
            conn = MySQLdb.connect(
                db='user_grade',
                host='localhost',
                user='root',
                password='123456',
                charset='utf8',
                unix_socket='/Applications/MAMP/tmp/mysql/mysql.sock'
            )
        except:
            pass
        return conn

    def export_data(self):

        conn = self.get_conn()
        cursor = conn.cursor()
        cursor.execute('SELECT `year`, `max_score`, `avg_score` FROM `user_score`')
        rows = cursor.fetchall()

        row_id = 10
        for (i, row) in enumerate(rows):
            (self.ws['C{0}'.format(row_id)],
             self.ws['D{0}'.format(row_id)],
             self.ws['E{0}'.format(row_id)]) = row
            row_id += 1

        chart = AreaChart()
        chart.title = '统计表'
        chart.style = 13
        chart.x_axis.title = '年份'
        chart.y_axis.title = '分数'

        cats = Reference(self.ws, min_col=3, min_row=10, max_row=row_id)
        data = Reference(self.ws, min_col=4, min_row=9, max_col=5, max_row=row_id)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        self.ws.add_chart(chart, "A{0}".format(row_id + 2))

        self.wb.save('./static/stats.xlsx')


if __name__ == '__main__':
    client = GaokaoExport()
    client.export_data()
